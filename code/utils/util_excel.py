import numpy as np
import openpyxl
import os

def load_excel_dataset(path='data/handwriting_dataset.xlsx', sheet_name='Data'):
    """
    이 함수는 6×6 픽셀로 구성된 숫자 이미지를 PyTorch에 바로 넣기 좋은 (96,1,6,6) 형태의 numpy 배열과,
    정답을 정수 레이블(0,1,2) 배열로 반환합니다.

    - 입력 블록: 각 숫자는 6행 × 6열 셀로 구성됩니다.
      첫 번째 샘플은 L열(12)부터 Q열(17)까지, 5~10행에 위치합니다.
      두 번째 샘플은 R열(18)부터 W열(23)까지, 5~10행에 위치하며,
      이후 6열씩 오른쪽으로 이동하여 총 96개 샘플이 배치되어 있습니다.

    - 정답 블록: 입력 블록 바로 아래의 3개 행(11~13행)에 one-hot 형태로 저장되어 있습니다.
      예를 들어, 첫 번째 샘플의 정답은 L11, L12, L13에 [1,0,0]으로,
      두 번째 샘플의 정답은 R11, R12, R13에 [0,1,0]으로,
      세 번째 샘플의 정답은 W11, W12, W13에 [0,0,1]으로 기록되어 있습니다.

    반환:
        X: numpy 배열, shape (96, 1, 6, 6)
        Y: numpy 배열, shape (96,), 값은 0, 1, 또는 2 (각 각 숫자 1, 2, 3에 대응)
    """

    if not os.path.exists(path):
        raise FileNotFoundError(f"Dataset not found at path: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[sheet_name]

    X_list = []
    Y_list = []

    for i in range(96):  # 96개의 샘플
        col_start = 12 + i * 6  # 첫 번째 샘플: L열(12), 두 번째: R열(18), …

        # 1) 입력: 5~10행 (6행) × col_start~col_start+5 (6열) → 6×6 2D 배열
        img_block = []
        for row in range(5, 11):  # 행 5~10
            row_vals = []
            for col in range(col_start, col_start + 6):  # 6열
                val = sheet.cell(row=row, column=col).value
                try:
                    row_vals.append(float(val))
                except (ValueError, TypeError):
                    row_vals.append(0.0)
            img_block.append(row_vals)
        X_list.append(img_block)

        # 2) 정답: 11~13행(3개)에서 한 열(col_start)만 가져오기 → one-hot 형태
        one_hot = []
        for row in range(11, 14):  # 행 11~13
            val = sheet.cell(row=row, column=col_start).value
            try:
                one_hot.append(int(val))
            except (ValueError, TypeError):
                one_hot.append(0)
        # one_hot 예: [1,0,0] 또는 [0,1,0] 또는 [0,0,1]
        # 0번째 인덱스가 1이면 레이블 0, 1번째 인덱스가 1이면 레이블 1, 2번째 인덱스가 1이면 레이블 2
        label = int(np.argmax(one_hot))
        Y_list.append(label)

    # X: (96, 6, 6) → (96, 1, 6, 6)
    X = np.array(X_list, dtype=np.float32).reshape(-1, 1, 6, 6)
    # Y: (96,)
    Y = np.array(Y_list, dtype=np.int64)

    return X, Y

# 테스트 블록
if __name__ == "__main__":
    X, Y = load_excel_dataset()
    print("✅ 전체 데이터 로딩 완료")
    print("X shape:", X.shape)  # (96, 1, 6, 6)
    print("Y shape:", Y.shape)  # (96,)
    print("첫 번째 입력 (1×6×6 배열):\n", X[0])
    print("첫 번째 정답 (정수 레이블):", Y[0])  # 0, 1, 또는 2
