import numpy as np
import openpyxl
import os

def load_excel_dataset(path='data/handwriting_dataset.xlsx', sheet_name='학습'):
    """
    이 함수는 6×6 픽셀로 구성된 숫자 이미지를 (flatten된) 36차원 벡터로 읽어옵니다.
    CNN에서는 입력을 (1, 6, 6) 형태로 다시 reshape하여 사용합니다.

    - 입력 블록: 각 숫자는 6행 × 6열 셀로 구성되며, flatten하여 36차원 벡터로 변환됩니다.
    - 정답 블록: 바로 아래 3개의 셀로 one-hot 형태의 정답 레이블이 들어 있습니다.
    - 총 96개의 데이터가 오른쪽으로 6칸씩 이동하며 구성되어 있습니다.

    예시:
        입력 36차원 벡터 (X[i]) → reshape → (1, 6, 6) 이미지 형태로 사용 가능
        정답 one-hot 벡터 (Y[i]) → [1,0,0], [0,1,0], [0,0,1] 중 하나
    """

    if not os.path.exists(path):
        raise FileNotFoundError(f"Dataset not found at path: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[sheet_name]

    X_list = []
    Y_list = []

    for i in range(96):  # 96개의 입력 블록
        col_start = 11 + i * 6  # K열(11)부터 시작해서 6칸씩 이동
        input_block = []
        
        # 입력: 6행 × 6열 (3~8행)
        for row in range(3, 9):  # 행 3~8
            for col in range(col_start, col_start + 6):  # 6열
                val = sheet.cell(row=row, column=col).value
                try:
                    input_block.append(float(val))
                except (ValueError, TypeError):
                    input_block.append(0.0)
        X_list.append(input_block)

        # 정답: 아래 3칸 (9~11행, 시작 열만 사용)
        label_block = []
        for row in range(9, 12):
            val = sheet.cell(row=row, column=col_start).value
            try:
                label_block.append(float(val))
            except (ValueError, TypeError):
                label_block.append(0.0)
        Y_list.append(label_block)

    X = np.array(X_list, dtype=np.float32)
    Y = np.array(Y_list, dtype=np.float32)

    return X, Y

# 테스트
if __name__ == "__main__":
    X, Y = load_excel_dataset()
    print("✅ 전체 데이터 로딩 완료")
    print("X shape:", X.shape)  # (96, 36)
    print("Y shape:", Y.shape)  # (96, 3)
    print("첫 번째 입력:", X[95])
    print("첫 번째 정답:", Y[95])
